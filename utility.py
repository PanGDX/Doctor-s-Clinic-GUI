import openpyxl,os, json
from datetime import datetime
def find_file(file_name):
    """
    Search through the root_folder and all its subfolders for a file named file_name.
    Returns a list of full paths to the file if found, otherwise returns False.
    """
    matches = []
    for root, dirs, files in os.walk(os.getcwd()):
        if file_name in files:
            matches.append(os.path.join(root, file_name))
    try:
        return matches[0]
    except:
        return False
    # must include the extension
    # returns full path

def make_file(dir, name):
    if(find_file(name) == False):
        print(f"Making file: {name}")
        with open(f"{dir}\\{name}", 'w') as file:
            if ".json" in name:
                json.dump({}, file, indent=4)
            elif ".xlsx" in name:
                file.close()
                wb = openpyxl.Workbook()
                #wb.create_sheet("dummy")
                wb.save(f"{dir}\\{name}")
            else:
                pass  # Creates an empty file

def make_xlsx(dir, name, list_to_append:list = []):
    if(find_file(name) == False):
        print(f"Making file: {name}")
        wb = openpyxl.Workbook()
        wb.active.append(list_to_append)
        wb.save(f"{dir}\\{name}")


def remove_blank_rows(excel_file):
    """
    removes all blank rows and saves the file
    """

    # Load the workbook and select the active worksheet
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    # Iterate backwards through the rows
    for row in range(sheet.max_row, 0, -1):
        if all(
            sheet.cell(row=row, column=col).value is None
            for col in range(1, sheet.max_column + 1)
        ):
            sheet.delete_rows(row)

    # Save the modified workbook
    workbook.save(excel_file)
    workbook.close()


def time_difference(timein: str, timeout: str):
    """
    calculates the time difference using string formatting, 24 hours
    """
    timeinlist = timein.split(":")
    timeoutlist = timeout.split(":")

    hourdifference = int(timeoutlist[0]) - int(timeinlist[0])
    minutedifference = int(timeoutlist[1]) - int(timeinlist[1])
    return hourdifference * 60 + minutedifference

def json_loading(file_name):
    """
    searches and loads the json file or else return empty json
    """
    file_path = find_file(file_name)
    if file_path:
        textfile = open(file_path, "r")

        json_file = json.loads(textfile.read())
        return json_file
    else:
        return {}
    

def load_and_check():
    """
    checks and creates all the necessary files
    """
    directory_name = "Files"
    if not os.path.exists(directory_name):
        print(f"Making Files folder")
        os.mkdir(directory_name)

    make_xlsx(f"{os.getcwd()}\\Files", "staff-log.xlsx", ["Date", "Name", "Time In", "Time Out", "Patients Treated"])
    make_xlsx(f"{os.getcwd()}\\Files", "medicine-log.xlsx")

    make_file(f"{os.getcwd()}\\Files", "medicine-treatment price.json")
    make_file(f"{os.getcwd()}\\Files", "staff-roles.json")
    make_file(f"{os.getcwd()}\\Files", "staff-wages.json")
    make_file(f"{os.getcwd()}\\Files", "barcode.json")

    if not os.path.exists(f"{os.getcwd()}\\Files\\patient info"):
        print(f"Making 'patient info' folder")
        os.mkdir(f"{os.getcwd()}\\Files\\patient info")

    make_file(f"{os.getcwd()}\\Files\\patient info", "name to id.json")

    if find_file('app.exe') == False:
        print("Missing app.exe (for staff)")
    if find_file('calculation.exe') == False:
        print("Missing calculation.exe (for owner(s))")

























def get_name_from_json(id: str):
    """
    get name from ID in json file: name to id.json

    if nothing is found, returns False

    WORKS
    """
    name_to_id_json = json_loading("name to id.json")
    for key in name_to_id_json:
        if name_to_id_json[key] == id:
            return key
    return False

def log_name_to_id(name: str, id: str):
    """
    log name to ID in json file: name to id.json

    WORKS
    """
    name_to_id_json = json_loading("name to id.json")
    exists = False
    for key in name_to_id_json:
        if key == name:
            exists = True
    if not exists:
        name_to_id_json[name] = id

        file_opened = open(find_file("name to id.json"), "w")
        json.dump(name_to_id_json, file_opened, indent=4)
        file_opened.close()



def create_medicine_sheet():
    """
    creates a new sheet if name does not exist
    If it does exist, ignore

    file does always exist (presumed)

    WORKS
    """

    excel_location = find_file("medicine-log.xlsx")
    workbook = openpyxl.load_workbook(excel_location)
    medicine_treatment_list = json_loading("medicine-treatment price.json")
    for key in medicine_treatment_list:
        if key not in workbook.sheetnames:
            workbook.create_sheet(key)
            worksheet = workbook[key]
            worksheet.append(["Date", "Quantity"])
    
    workbook.save(excel_location)



def calculate_costs(quantity_json):
    """
    input json file
    reads json file
    calculates the cost

    specifically made for date. eliminates any (type):"str" elements from calculation
    """
    total_cost = 0
    pricing_json = json_loading("medicine-treatment price.json")
    for key in quantity_json:
        if type(quantity_json[key]) == int:
            total_cost += (pricing_json[key][0] * quantity_json[key])

    return total_cost


def modify_medicine_log(medicine_treatment_name = "",quantity = 0):
    """
    assumes existence of sheet in excel file
    for current date only
    if current date exist, add/subtract
    note that it does NOT STACK -> add/subtract separated!

    quantity can be negative!!
    
    if current date does not exist, add new entry. 
    KEY:sign included

    WORKS
    """

    excel_location = find_file("medicine-log.xlsx")
    now = datetime.now()
    date = now.strftime("%d/%m/%Y")

    workbook = openpyxl.load_workbook(excel_location)
    worksheet = workbook[f"{medicine_treatment_name}"]
    
    modified = False
    for i, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == date: #date exists
            if row[1] != None:
                if row[1] < 0 and quantity<0:
                    modified=True
                    #index 1 for cell modification
                    worksheet.cell(row = i, column=2, value= row[1] + quantity)
                if row[1]>0 and quantity>0:
                    modified=True
                    #index 1 for cell modification
                    worksheet.cell(row = i, column=2, value= row[1] + quantity)
    if not modified:
        worksheet.append([date, quantity])
        

    workbook.save(excel_location)


def save_to_patient_file(patient_ID: str, medicine_or_treatment = "",  quantity=0, doctor_note=""):
    """
    save to patient file in Files\\patient info

    do not need to log the name of the patient. This information can be obtained from json file


    medicine = "", treatment="", quantity=0, doctor_note=""
    fill in the optional information. MUST DO
    """


    
    make_file(f"{os.getcwd()}\\Files\\patient info", f"{patient_ID}.json")

    #error no saving

    output_json = json_loading(f"{patient_ID}.json")
    print(output_json)
    output_json["ID"] = patient_ID 
    # somewhat redundant but whatever


    # Get the current date and time
    now = datetime.now()
    date = now.strftime("%d/%m/%Y")
    
    if output_json.get(date) == None:
        
        #new date
        output_json[date] = {}

        if(medicine_or_treatment !=""):
            output_json[date][medicine_or_treatment] = quantity
        if(doctor_note != ""):
            output_json[date]["doctor note"] = doctor_note
    else:
        #append to that
        if(medicine_or_treatment !=""):
            if output_json[date].get(medicine_or_treatment) == None:
                output_json[date][medicine_or_treatment] = quantity
            else:
                output_json[date][medicine_or_treatment] += quantity
        if(doctor_note != ""):
            if output_json[date].get("doctor note") == None:
                output_json[date]["doctor note"] = doctor_note
            else:
                output_json[date]["doctor note"] = output_json[date]["doctor note"] + "\n"+ doctor_note
            

    
    print(output_json)


    file_opened = open(find_file(f"{patient_ID}.json"), "w")
    json.dump(output_json, file_opened, indent=4)
    file_opened.close()
 
 
