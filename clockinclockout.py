import tkinter as tk
from tkinter import messagebox
from datetime import datetime
from openpyxl import load_workbook
import os
from utility import find_file,remove_blank_rows,json_loading


#NOTE:
# modified to staff-roles, not reliant on staff names anymore


def clockinclockout(frame):
    # loads the file of all the names
    names = []
    for key in json_loading("staff-roles.json"):
        names.append(key)


    # Load or create the workbook
    log_location = find_file('staff-log.xlsx')
    remove_blank_rows(log_location)
    workbook = load_workbook(filename=log_location)


    # Create the main window and set its size
    input_var = tk.StringVar()
    
    message = tk.Message(
        frame,
        text ="Names",
        width=500,
    )
    # Create a listbox for the names and set the font size
    name_listbox = tk.Listbox(
        frame, listvariable=tk.Variable(value=names), width=50, height=20
    )
    name_listbox.pack(pady=20)



    # Function to handle clocking in
    def clock_in():
        # Check if a name is selected
        if not name_listbox.curselection():
            messagebox.showinfo("Clock In", "ไม่ได้เลือกชื่อ")
            
            return
        # Get the selected name
        name = name_listbox.get(name_listbox.curselection())

        # print(name)
        for i, row in enumerate(
            workbook.active.iter_rows(min_row=2, values_only=True), start=2
        ):
            # print(row)
            if row[1] == name:
                if row[3] == "" or row[3] == None:
                    messagebox.showinfo(
                        "Clock In", f"{name}ได้เข้าสู่ระบบแล้วและยังไม่ได้ออกจากระบบ"
                    )
                    return 
        # code that checks if the person has clocked in on this day

        # Get the current date and time
        now = datetime.now()
        date = now.strftime("%d/%m/%Y")

        if now.minute < 10:
            time = f"{now.hour}:0{now.minute}"
        else:
            time = f"{now.hour}:{now.minute}"
        # Show a message box

        messagebox.showinfo("Clock In", f"{name} ได้เข้าสู่ระบบ: {date} {time}")

        # Write to the workbook
        workbook.active.append([date, name, time, ""])
        workbook.save(filename=f"{os.getcwd()}\\Files\\staff-log.xlsx")


    # Function to handle clocking out
    def clock_out():
        input_str = input_var.get()

        # Check if a name is selected
        if not name_listbox.curselection():
            messagebox.showinfo("Clock Out", "ไม่ได้เลือกชื่อ")
            return
        # check if number of treated patients is filled
        if not input_str:
            messagebox.showinfo(
                "Clock Out", "ไม่มีการป้อนข้อมูลหมายเลขผู้ป่วย (Patient Input)"
            )
            return

        # check if input field is an integer
        try:
            int(input_str)
        except:
            messagebox.showinfo("Clock Out", "ไม่ใช่จำนวน")
            return

        # Get the selected name
        name = name_listbox.get(name_listbox.curselection())

        # Get the current date and time
        now = datetime.now()
        date = now.strftime("%d/%m/%Y")
        if now.minute < 10:
            time = f"{now.hour}:0{now.minute}"
        else:
            time = f"{now.hour}:{now.minute}"
        # Check if the user has clocked in
        for i, row in enumerate(
            workbook.active.iter_rows(min_row=2, values_only=True), start=2
        ):
            if (
                now.strftime("%d/%m/%Y") == row[0]
                and row[1] == name
                and row[2]
                and not row[3]
            ):
                # Update the clock out time

                workbook.active.cell(row=i, column=4, value=time)
                workbook.active.cell(row=i, column=5, value=int(input_str))
                workbook.save(filename=f"{os.getcwd()}\\Files\\staff-log.xlsx")
                # Show a message box
                messagebox.showinfo("Clock Out", f"{name} ได้ออกจากระบบ: {date} {time}")

                return
        # If the user has not clocked in, show a message box


    # Create the clock in button
    clock_in_button = tk.Button(
        frame, text="Clock In", command=clock_in, width=8, height=1
    )
    clock_in_button.pack()

    # Create the clock out button
    clock_out_button = tk.Button(
        frame, text="Clock Out", command=clock_out, width= 8, height=1
    )
    clock_out_button.pack(pady=20)

    message = tk.Message(
        frame,
        text="Input the number of patients:",
        width=500,
        justify="center",
    )
    message.pack(pady=10)

    entry = tk.Entry(frame, textvariable=input_var, width=50)
    entry.pack(pady=20)


    message = tk.Message(
        frame,
        text="To edit the name list, go to settings and choose 'Staff Names'",
        width=500,
    )
    message.pack()

    # Start the main loop
    

if __name__ == '__main__':
    ws = tk.Tk()
    clockinclockout(ws)
    ws.mainloop()  # This line starts the Tkinter event loop
