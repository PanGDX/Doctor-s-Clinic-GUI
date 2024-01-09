import openpyxl
from utility import *
import math, os, sys
from datetime import datetime
from tkinter import *
from tkcalendar import Calendar
import time







global start_datetime_object, end_datetime_object






def delete_sheet(file_path, sheet_name):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Check if the sheet exists in the workbook
    if sheet_name in workbook.sheetnames:
        # Remove the sheet
        workbook.remove(workbook[sheet_name])
        print(f"Sheet '{sheet_name}' has been deleted.")

        # Save the workbook
        workbook.save(file_path)
        print(f"Changes saved to '{file_path}'.")
    else:
        print(f"Sheet '{sheet_name}' does not exist in the workbook.")
    workbook.close()


def save_staff_log_to_excel(sql_connection, excel_file_path):
    # Connect to the MySQL database
    if find_file("staff-log.xlsx") != False:
        os.remove(find_file("staff-log.xlsx"))
        make_file(os.getcwd(),'staff-log.xlsx',["Date", "Name", "TimeIn", "TimeOut", "Patients"])
        print("Processing")
        time.sleep(5)

    with sql_connection.cursor() as cursor:
        # SQL query to select data from Clinic.StaffLog
        sql = "SELECT Date,Name,TimeIn,TimeOut,Patients FROM Clinic.StaffLog"
        cursor.execute(sql)

        # Fetch all rows from the database
        rows = cursor.fetchall()

        # Create an Excel workbook and sheet
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Writing the data
        for row in rows:
            sheet.append(list(row.values()))

        # Save the workbook to the specified file path
        workbook.save(excel_file_path)
        workbook.close()


def save_medicine_log_to_excel(sql_connection, excel_file_path):
    # Connect to the MySQL database

    if find_file("medicine-log.xlsx") != False:
        os.remove(find_file("medicine-log.xlsx"))
        create_medicine_sheet(sql_connection)
        print("Processing")
        time.sleep(5)
        
    # Load the existing Excel workbook
    workbook = openpyxl.load_workbook(excel_file_path)

    with sql_connection.cursor() as cursor:
        # SQL query to select data from the Medicine Log
        sql = "SELECT Date, Name, Quantity FROM MedicineLog"
        cursor.execute(sql)

        # Fetch all rows from the database
        rows = cursor.fetchall()

        for row in rows:
            # Determine the sheet name from the medicine name
            sheet_name = row['Name']

            # Select the sheet
            sheet = workbook[sheet_name]

            # Append the Date and Quantity to the sheet
            sheet.append([row['Date'], row['Quantity']])

        # Save the updated workbook
        workbook.save(excel_file_path)
        workbook.close()







def select_dates():
    global start_datetime_object, end_datetime_object

    def grad_date():
        global start_datetime_object, end_datetime_object
        start_datetime_object = datetime.strptime(calstart.get_date(), "%m/%d/%y")
        end_datetime_object = datetime.strptime(calend.get_date(), "%m/%d/%y")

        formatted_start_date = start_datetime_object.strftime("%d/%m/%Y")
        formatted_end_date = end_datetime_object.strftime("%d/%m/%Y")
        print(
            f"Date selected (inclusive): {formatted_start_date} to {formatted_end_date}"
        )

        root.destroy()

    root = Tk()
    root.geometry("600x700")

    label1 = Label(root, text="First date to include\n", font=("default", 14))
    label1.pack(pady=10)
    calstart = Calendar(
        root,
        selectmode="day",
        year=datetime.now().year,
        month=datetime.now().month,
        day=datetime.now().day,
    )
    calstart.pack(pady=20)

    label2 = Label(root, text="Last date to include\n", font=("default", 14))
    label2.pack(pady=10)
    calend = Calendar(
        root,
        selectmode="day",
        year=datetime.now().year,
        month=datetime.now().month,
        day=datetime.now().day,
    )
    calend.pack(pady=20)

    Button(root, text="Get Date", command=grad_date).pack(pady=20)
    root.attributes("-topmost", True)
    root.mainloop()


def merge_function():
    workbook = openpyxl.load_workbook(filename=find_file('staff-log.xlsx'))
    sheet = workbook.active

    data = []
    missing_fields = False
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        date, name, clock_in, clock_out, patients = row
        data.append(
            {
                "Date": date,
                "Name": name,
                "Clock In": clock_in,
                "Clock Out": clock_out,
                "Patients": patients,
            }
        )
        if clock_out == None or clock_out == "":
            print(f"Missing clock out time on {date} by {name}")
            missing_fields = True

    if missing_fields:
        input("Please fill in the missing fields. Input anything to exit.")
        sys.exit()

    merged_data = {}
    for row in data:
        key = (row["Date"], row["Name"])
        if key not in merged_data:
            merged_data[key] = {
                "Clock In": row["Clock In"],
                "Clock Out": row["Clock Out"],
                "Patients": row["Patients"],
            }
        else:
            merged_data[key]["Clock Out"] = row["Clock Out"]
            merged_data[key]["Patients"] += row["Patients"]

    mergedworkbook = openpyxl.Workbook()
    

    mergedworkbook.active.append(
        ["Date", "Name", "Clock In", "Clock Out", "Patients"]
    )  # Header


    for key, row in merged_data.items():
        if type(key[0]) == str:
            date = key[0]
        else:   
            date = key[0].strftime("%m/%d/%Y")
            print(type(key[0]))
            print(date)
        #format is d/t/m
        # but if you type in 3/2/2023 -> usually excel reads as month 3, day 2

        mergedworkbook.active.append(
            [date, key[1], row["Clock In"], row["Clock Out"], row["Patients"]]
        )

    mergedworkbook.save("staff-log.xlsx")


# Main Function for staff-log processing
def TimeWorked(sql_connection):
    def WagePerDay(minutes: int, types: str, date: int, patients: int, connection):
        

        # READ ROLES WAGES FROM SQL
        cursor = connection.cursor()
        query = "SELECT * FROM Clinic.RolesPayment"
        cursor.execute(query)
        results = cursor.fetchall()
        cursor.close()

        roles = {}
        for key_dict in results:
            roles[key_dict['Roles']] = {
                "WeekdayDollar":key_dict["WeekdayDollar"],
                "WeekendDollar":key_dict["WeekendDollar"],
                "OvertimeWeekdayDollar":key_dict["OvertimeWeekdayDollar"],
                "OvertimeWeekendDollar":key_dict["OvertimeWeekendDollar"]
            }

        summation = 0
        summation += 40
        summation += patients


        if patients >= 30:
            summation += 50

        if date < 6:  # weekdays
            time_remaining = max(minutes - 3 * 60, 0) / 30

            summation += roles[types]["WeekdayDollar"]
            summation += (roles[types]["OvertimeWeekdayDollar"] * math.floor(time_remaining))

        if date >= 6:  # weekends
            time_remaining = max(minutes - 4 * 60, 0) / 30

            summation += roles[types]["WeekendDollar"]
            summation += (roles[types]["OvertimeWeekendDollary"] * math.floor(time_remaining))

        return summation

    dictionary_of_names_details = {}

    workbook = openpyxl.load_workbook(filename=find_file("staff-log.xlsx"))

    
    # READ ROLES FROM SQL
    list_of_staff_roles = query_staff_roles(sql_connection)
    for key_dict in list_of_staff_roles:
        dictionary_of_names_details[key_dict['Name']] = [key_dict['Role'], 0, 0, 0, 0, 0]



    input("Input anything to continue")
    os.system("cls")

    print("Select the date range to include in the calculation")
    select_dates()
    for i, row in enumerate(workbook.active.iter_rows(min_row=2, values_only=True), start=2):
        date = row[0]
        try:
            datetime_object = datetime.strptime(date, "%d/%m/%Y")
            # print(date)
        except:
            date_now = f"{date.month}/{date.day}/{date.year}"
            # print(date_now)
            datetime_object = datetime.strptime(date_now, "%d/%m/%Y")

        minutes = 0
        refer_date = None
        patients = 0
        paid = 0
        if (start_datetime_object <= datetime_object and datetime_object <= end_datetime_object):
            minutes = time_difference(row[2], row[3])
            refer_date = (datetime_object.weekday() + 1)  
            # 1 for monday, 7 for sunday, etc
            

            patients = int(row[4])

            paid = WagePerDay(
                minutes, dictionary_of_names_details[row[1]][0], refer_date, patients, sql_connection
            )

        
        if refer_date != None:
            dictionary_of_names_details[row[1]][1] += paid
            
            if refer_date < 6:
                dictionary_of_names_details[row[1]][2] += 1
                dictionary_of_names_details[row[1]][4] += round(minutes / 60.0, 2)
            if refer_date >= 6:
                dictionary_of_names_details[row[1]][3] += 1
                dictionary_of_names_details[row[1]][5] += round(minutes / 60.0, 2)
        else:
            print("Some error somewhere??")

    for x, y in dictionary_of_names_details.items():
        if y[1] != 0:
            if y[0] == "Helper":
                print(
                    f"{x} (helper) should be paid {y[1]} baht.\nThey worked for {y[3]} weekends for {y[5]} hours.\nThey worked for {y[2]} weekdays for {y[4]} hours\n\n"
                )

            elif y[0] == "Nurse":
                print(
                    f"{x} (nurse) should be paid {y[1]} baht.\nThey worked for {y[3]} weekends for {y[5]} hours.\nThey worked for {y[2]} weekdays for {y[4]} hours\n\n"
                )

            else:
                print(
                    f"{x} should be paid {y[1]} baht.\nThey worked for {y[3]} weekends for {y[5]} hours.\nThey worked for {y[2]} weekdays for {y[4]} hours\n\n"
                )
    input("Input anything to continue")
    sys.exit()

def MedicineCalculation(sql_connection):
    delete_sheet(find_file('medicine-log.xlsx'), 'Sheet')
    """
    - read the xlsx data
    - get the information on price from sql
    - collate and calculate into a dictionary. Formatted:
    {"Name": {
        "Leftover":
        "Used":
        }
    }

    and then we can use information from sql to calculate
    """
    final_data = {}
    total_revenue = 0
    total_profit = 0
    total_cost = 0

    
    workbook = openpyxl.load_workbook(filename=find_file('medicine-log.xlsx'))
    select_dates()
    for sheet in workbook.sheetnames:
        current_sheet = workbook[sheet]
        

        final_data[sheet] = {"Leftover":0, 
                             "Used":0}
    
        for i, row in enumerate(current_sheet.iter_rows(min_row=2, values_only=True), start=2):
            date = row[0]
            try:
                datetime_object = datetime.strptime(date, "%d/%m/%Y")
                # print(date)
            except:
                date_now = f"{date.month}/{date.day}/{date.year}"
                # print(date_now)
                datetime_object = datetime.strptime(date_now, "%d/%m/%Y")
            quantity = int(row[1])

            if (start_datetime_object <= datetime_object and datetime_object <= end_datetime_object):
                if quantity < 0:
                    final_data[sheet]["Used"] = final_data[sheet]["Used"] + abs(quantity)
                
                final_data[sheet]["Leftover"] = final_data[sheet]["Leftover"] + quantity

    cursor = sql_connection.cursor()
    query = "SELECT * FROM Clinic.MedicineTreatmentPrices"
    cursor.execute(query)
    result = cursor.fetchall()
    cursor.close()

    medicine_data_in_dict = {}
    for key_dict in result:
        medicine_data_in_dict[key_dict['Name']] = {
            "Price":key_dict["Price"],
            "CostPrice":key_dict['CostPrice']
        }

    for name in final_data:
        leftover_quantity = final_data[name]["Leftover"]
        used_quantity = final_data[name]["Used"]
        revenue_per_unit = medicine_data_in_dict[name]["Price"]
        cost_per_unit = medicine_data_in_dict[name]["CostPrice"]
        profit_per_unit = revenue_per_unit - cost_per_unit

        print(f"For {name}:")
        print(f"Stock leftover: {leftover_quantity}")
        print(f"Used: {used_quantity}")
        print(f"Revenue: {used_quantity * revenue_per_unit}")
        print(f"Profit: {used_quantity * profit_per_unit}")
        print(f"Cost spent: {used_quantity * cost_per_unit}")


        total_revenue += (used_quantity * revenue_per_unit)
        total_profit += (used_quantity * profit_per_unit)
        total_cost += (used_quantity * cost_per_unit)

    print("\n\n\n")
    print(f"Total revenue: {total_revenue}")
    print(f"Total cost: {total_cost}")
    print(f"Total profit: {total_profit}")

if __name__ == "__main__":
    """
    How this program works:
    - Checks for 2 xlsx files - staff-log and medicine-log (DONE)
    - Creates them if they do not exist (function written) (DONE)
    - Pulls data from sql to xlsx files (modify and check) (OPTION GIVEN)
    - Merge the data for the staff-log. For the medicine-log, no need to do so.

    - Select options: 
        - View staff-log
            - Gives data. Already written
        - View medicine-log 
            - What data to calculate?
            - differentiate between + and -
            - how many used
            - profit, revenue, cost (from usage)
            - stock leftover (first input will be +, starting amount)

    """
    sql_connection = return_connection()


    
    option = int(input("Input 1: Load excel data from database. Will delete the file! Warning!\nInput 2: Skip this step, use existing excel data\n:"))
    if option == 1:
        # load data
        save_staff_log_to_excel(sql_connection,find_file('staff-log.xlsx'))
        save_medicine_log_to_excel(sql_connection, find_file('medicine-log.xlsx'))

    option = int(input("Input 1: Calculate data for staff\nInput 2: Calculate data for medicine\n:"))
    if option == 1:
        merge_function()
        # merge staff-log
        # run calculations
        TimeWorked(sql_connection)
    if option == 2:
        # process medicine-data
        MedicineCalculation(sql_connection)

