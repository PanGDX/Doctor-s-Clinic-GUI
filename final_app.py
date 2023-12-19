import tkinter as tk 
import tkinter.ttk as ttk
from clockinclockout import clockinclockout
from openpyxl import load_workbook
from utility import find_file
from settings import setup_json_editor
from medicine import medicine
from patient_info_request import patient_info
import sys,asyncio
from threading import Thread


# remaining things to do: test async with actual barcode scanner
# test inputs


def on_closing():
    """Function to call when window is closing."""

    loop.call_soon_threadsafe(loop.stop)
    t.join()

    log_location = find_file('staff-log.xlsx')
    workbook = load_workbook(filename=log_location)
    empty = False
    namelist = []
    for i, row in enumerate(
            workbook.active.iter_rows(min_row=2, values_only=True), start=2
        ):
            # print(row)
            if row[3] == "" or row[3] == None:
                empty = True
                namelist.append(row[1])



    if empty:
        tk.messagebox.showwarning("Warning", f"Somebody did not log out yet.\nNames: {' '.join(namelist)}")
    else:
        ws.destroy()

# Create the main window
ws = tk.Tk()
ws.geometry('1000x700')
ws.title('Notebook')



# Create a Notebook widget
notebook = ttk.Notebook(ws)
style = ttk.Style()
 

# COLOR_BLUE = "#89CFF0"

# style.theme_create("yummy", settings={
#     "TNotebook.Tab": {
#         "map": {
#             "background": [("selected", COLOR_BLUE)],   
#             "font": [("selected", ('bold'))]    
#             } 
#         } 
#     } 
# )

# style.theme_use("yummy")
#Issue why the fuck does the size expands??

# Create four frames
staff_frame = ttk.Frame(notebook)
medicine_frame = ttk.Frame(notebook)
patient_info_frame = ttk.Frame(notebook)
settings_frame = ttk.Frame(notebook)
# Add a label to each frame


clockinclockout(staff_frame)
setup_json_editor(ws, settings_frame)
loop,t=medicine(medicine_frame)
patient_info(patient_info_frame)
# Pack the frames
staff_frame.pack(fill='both', expand=True)
medicine_frame.pack(fill='both', expand=True)


# Add frames to the notebook
notebook.add(staff_frame, text='  Staff Login  ')
notebook.add(medicine_frame, text= '  Medicine and Treatment Logging  ') # similar to staff login
notebook.add(patient_info_frame, text= '  Search Patient Information  ') # canvas?
notebook.add(settings_frame, text= '  Settings  ') #treeview
notebook.pack(padx=10, pady=10, expand=True, fill='both')

# Run the main event loop
ws.protocol("WM_DELETE_WINDOW", on_closing)
ws.mainloop()
